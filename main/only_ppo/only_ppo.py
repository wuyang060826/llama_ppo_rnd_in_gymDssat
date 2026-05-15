#!/usr/bin/env python3
"""
Pure PPO Baseline for Gym-DSSAT Crop Management
================================================
A strong baseline for agricultural reinforcement learning research.

Features:
- Pure PPO implementation (no LLM, numerical state input only)
- Improved A2C family algorithm with excellent stability
- Better sample efficiency than DQN
- Comprehensive agricultural and AI metrics tracking (strictly defined)
- Optimized for A100 40G GPU

References:
- Schulman et al. (2017) "Proximal Policy Optimization Algorithms"
- Schulman et al. (2016) "High-Dimensional Continuous Control Using GAE"
- Mnih et al. (2016) "Asynchronous Methods for Deep Reinforcement Learning"
"""

import numpy as np
import pandas as pd
import random
from collections import deque, OrderedDict
import time
import math
import json
import os
from tqdm import tqdm
from dataclasses import dataclass, field
from typing import List, Tuple, Optional, Dict, Any
import warnings
warnings.filterwarnings('ignore')

import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torch.distributions import Categorical
from torch.utils.data import DataLoader, TensorDataset

import gym

import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # 非交互式后端
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# 显存碎片优化
os.environ["PYTORCH_ALLOC_CONF"] = "expandable_segments:True"

# 结果保存目录
RESULTS_DIR = '/home/wuyang/results/only_ppo_results/0410'
os.makedirs(RESULTS_DIR, exist_ok=True)

# ============================================================================
#                              设备配置
# ============================================================================

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print("=" * 70)
print("🚀 Pure PPO Baseline - Gym-DSSAT Crop Management")
print("=" * 70)
print(f"📍 使用设备: {device}")
if torch.cuda.is_available():
    print(f"   GPU型号: {torch.cuda.get_device_name(0)}")
    print(f"   显存总量: {torch.cuda.get_device_properties(0).total_memory / 1024**3:.1f} GB")
    # 开启TF32加速 (A100支持)
    torch.backends.cuda.matmul.allow_tf32 = True
    torch.backends.cudnn.allow_tf32 = True
    print(f"   TF32加速: 已开启")


# ============================================================================
#                              超参数配置
# ============================================================================

@dataclass
class PPOConfig:
    """PPO超参数配置 - 纯数值输入版本
    
    作为论文强基线，参数设置参考：
    - Stable Baselines3 默认参数
    - RLlib PPO 推荐配置
    - 相关农业RL论文实践
    """
    
    # === 训练参数 ===
    n_episodes: int = 3000              # 训练轮数
    max_steps_per_episode: int = 200    # 每轮最大步数
    
    # === PPO核心参数 ===
    gamma: float = 0.99                 # 折扣因子
    gae_lambda: float = 0.95            # GAE lambda参数
    clip_ratio: float = 0.2             # PPO裁剪比例
    target_kl: float = 0.02             # KL散度目标值
    entropy_coef: float = 0.01          # 熵正则化系数
    value_coef: float = 0.5             # 价值损失系数
    max_grad_norm: float = 0.5          # 梯度裁剪阈值
    
    # === 优化器参数 ===
    learning_rate: float = 3e-4         # 统一学习率
    weight_decay: float = 1e-4          # 权重衰减
    betas: Tuple[float, float] = (0.9, 0.999)  # Adam betas
    eps: float = 1e-8                   # Adam epsilon
    
    # === 学习率调度 ===
    use_lr_schedule: bool = True        # 是否使用学习率调度
    warmup_steps: int = 100             # 预热步数
    min_lr_ratio: float = 0.1           # 最小学习率比例
    
    # === PPO更新参数 ===
    ppo_epochs: int = 10                # 每次更新的PPO轮数
    mini_batch_size: int = 64           # Mini-batch大小
    update_frequency: int = 10          # 更新频率（每N轮）
    
    # === 网络参数 ===
    state_size: int = 25                # 状态维度
    action_size: int = 25               # 动作维度 (5氮肥 × 5灌溉)
    hidden_sizes: List[int] = field(default_factory=lambda: [256, 256])  # 隐藏层大小
    
    # === 奖励函数参数 (根据原代码) ===
    k1: float = 0.158                   # 产量奖励系数
    k2: float = 0.79                    # 氮肥成本系数
    k3: float = 1.1                     # 灌溉成本系数
    
    # === 探索参数 ===
    entropy_decay: float = 0.999        # 熵系数衰减
    entropy_min: float = 0.001          # 最小熵系数
    
    # === 优化选项 ===
    use_bf16: bool = True               # A100使用BF16
    
    # === 验证参数 ===
    eval_frequency: int = 100           # 评估频率
    n_eval_episodes: int = 10           # 评估轮数
    
    # === 日志参数 ===
    log_frequency: int = 10             # 日志输出频率
    save_frequency: int = 1000           # 模型保存频率
    
    # === 指标参数 ===
    # 样本效率目标性能（专家水平阈值，基于奖励函数估算）
    # 假设专家产量约8000 kg/ha，施肥约120 kg N/ha，灌溉约100 mm
    # 专家奖励 ≈ k1 * 8000 - k2 * 120 - k3 * 100 = 1264 - 94.8 - 110 = 1059.2
    target_performance: float = 800.0   # 目标性能阈值（保守估计）
    
    # 收敛判定参数
    convergence_window: int = 100       # 收敛判定窗口大小
    convergence_threshold: float = 0.05 # 收敛判定阈值（变化率）

config = PPOConfig()


# ============================================================================
#                              辅助函数
# ============================================================================

def dict2array(state: dict) -> np.ndarray:
    """将gym-dssat字典状态转换为numpy数组"""
    if state is None:
        raise ValueError("状态不能为None")
    new_state = []
    for key in state.keys():
        if key != 'sw':
            new_state.append(state[key])
        else:
            new_state += list(state['sw'])
    return np.asarray(new_state, dtype=np.float32)


def get_reward(state: np.ndarray, n_action: float, w_action: float, 
               next_state: np.ndarray, done: bool, 
               k1: float, k2: float, k3: float) -> float:
    """计算奖励值
    
    奖励函数设计：
    - 终止时：奖励 = k1 * 产量 - k2 * 氮肥量 - k3 * 灌溉量
    - 非终止时：奖励 = -k2 * 氮肥量 - k3 * 灌溉量
    """
    if done:
        return k1 * state[4] - k2 * n_action - k3 * w_action
    return -k2 * n_action - k3 * w_action


def print_gpu_memory(prefix: str = ""):
    """打印GPU显存使用情况"""
    if torch.cuda.is_available():
        allocated = torch.cuda.memory_allocated() / 1024**3
        reserved = torch.cuda.memory_reserved() / 1024**3
        print(f"📊 {prefix} GPU显存: 已用={allocated:.2f}GB, 预留={reserved:.2f}GB")


def set_seed(seed: int):
    """设置随机种子以确保可复现性"""
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
# ============================================================================
#                           学习率调度器
# ============================================================================

class CosineAnnealingWarmup:
    """余弦退火学习率调度器（带预热）"""
    
    def __init__(self, optimizer, warmup_steps: int, total_steps: int, 
                 min_lr_ratio: float = 0.1):
        self.optimizer = optimizer
        self.warmup_steps = warmup_steps
        self.total_steps = total_steps
        self.min_lr_ratio = min_lr_ratio
        self.current_step = 0
        self.base_lrs = [group['lr'] for group in optimizer.param_groups]
        
    def step(self):
        """更新学习率"""
        self.current_step += 1
        for i, group in enumerate(self.optimizer.param_groups):
            base_lr = self.base_lrs[i]
            if self.current_step < self.warmup_steps:
                lr = base_lr * self.current_step / self.warmup_steps
            else:
                progress = (self.current_step - self.warmup_steps) / max(1, self.total_steps - self.warmup_steps)
                lr = base_lr * (self.min_lr_ratio + (1 - self.min_lr_ratio) * 
                               (1 + math.cos(math.pi * progress)) / 2)
            group['lr'] = lr
    
    def get_lr(self) -> List[float]:
        """获取当前学习率"""
        return [group['lr'] for group in self.optimizer.param_groups]


# ============================================================================
#                              网络定义
# ============================================================================

class ActorCriticNetwork(nn.Module):
    """Actor-Critic网络
    
    共享特征提取层，分离的策略头和价值头
    这是PPO作为A2C改进版的核心架构
    """
    
    def __init__(self, state_size: int, action_size: int, 
                 hidden_sizes: List[int] = [256, 256]):
        super().__init__()
        
        # 共享特征提取层
        shared_layers = []
        prev_size = state_size
        
        for hidden_size in hidden_sizes[:-1]:
            linear = nn.Linear(prev_size, hidden_size)
            nn.init.orthogonal_(linear.weight, gain=np.sqrt(2))
            nn.init.constant_(linear.bias, 0.0)
            shared_layers.extend([
                linear,
                nn.LayerNorm(hidden_size),
                nn.ReLU(),
            ])
            prev_size = hidden_size
        
        self.shared = nn.Sequential(*shared_layers)
        
        # 策略头（Actor）
        self.actor = nn.Sequential(
            nn.Linear(prev_size, hidden_sizes[-1]),
            nn.LayerNorm(hidden_sizes[-1]),
            nn.ReLU(),
            nn.Linear(hidden_sizes[-1], action_size)
        )
        nn.init.orthogonal_(self.actor[-1].weight, gain=0.01)
        nn.init.constant_(self.actor[-1].bias, 0.0)
        
        # 价值头（Critic）
        self.critic = nn.Sequential(
            nn.Linear(prev_size, hidden_sizes[-1]),
            nn.LayerNorm(hidden_sizes[-1]),
            nn.ReLU(),
            nn.Linear(hidden_sizes[-1], 1)
        )
        nn.init.orthogonal_(self.critic[-1].weight, gain=1.0)
        nn.init.constant_(self.critic[-1].bias, 0.0)
        
    def forward(self, x):
        """前向传播，返回动作logits和状态价值"""
        features = self.shared(x)
        logits = self.actor(features)
        value = self.critic(features)
        return logits, value
    
    def get_action_value(self, x):
        """获取动作分布和价值"""
        logits, value = self.forward(x)
        dist = Categorical(logits=logits)
        return dist, value


# ============================================================================
#                              PPO Buffer
# ============================================================================

@dataclass
class Transition:
    """单步转移数据"""
    state: np.ndarray
    action: int
    reward: float
    next_state: np.ndarray
    done: bool
    log_prob: float
    value: float
    n_action: float  # 氮肥量
    w_action: float  # 灌溉量


class RolloutBuffer:
    """PPO经验回放缓冲区"""
    
    def __init__(self, gamma: float, gae_lambda: float, device: torch.device):
        self.gamma = gamma
        self.gae_lambda = gae_lambda
        self.device = device
        self.transitions: List[Transition] = []
        
        # 运行统计量（用于奖励归一化）
        self.reward_mean = 0.0
        self.reward_std = 1.0
        self.reward_count = 0
        
    def add(self, state, action, reward, next_state, done, 
            log_prob, value, n_action, w_action):
        """添加一条转移数据"""
        self.transitions.append(Transition(
            state, action, reward, next_state, done, 
            log_prob, value, n_action, w_action
        ))
    
    def clear(self):
        """清空缓冲区"""
        self.transitions = []
    
    def __len__(self):
        return len(self.transitions)
    
    def compute_gae(self, last_value: float, use_reward_norm: bool = True):
        """计算GAE和回报"""
        n = len(self.transitions)
        
        rewards = np.array([t.reward for t in self.transitions])
        
        if use_reward_norm:
            batch_mean = rewards.mean()
            batch_var = rewards.var()
            batch_count = len(rewards)
            
            delta = batch_mean - self.reward_mean
            total_count = self.reward_count + batch_count
            self.reward_mean += delta * batch_count / total_count
            
            m2 = self.reward_var * self.reward_count + batch_var * batch_count + delta**2 * self.reward_count * batch_count / total_count
            self.reward_var = m2 / total_count if total_count > 0 else 1.0
            self.reward_count = total_count
            
            reward_std = np.sqrt(self.reward_var) + 1e-8
            norm_rewards = (rewards - self.reward_mean) / reward_std
        else:
            norm_rewards = rewards
        
        rewards_t = torch.tensor(norm_rewards, dtype=torch.float32, device=self.device)
        values_t = torch.tensor([t.value for t in self.transitions], dtype=torch.float32, device=self.device)
        dones_t = torch.tensor([t.done for t in self.transitions], dtype=torch.float32, device=self.device)
        
        advantages = torch.zeros(n, dtype=torch.float32, device=self.device)
        gae = 0.0
        
        for t in reversed(range(n)):
            if t == n - 1:
                next_value = last_value
                next_non_terminal = 1.0 - dones_t[t]
                if dones_t[t]:
                    next_value = 0.0
            else:
                next_value = values_t[t + 1]
                next_non_terminal = 1.0 - dones_t[t]
            
            delta = rewards_t[t] + self.gamma * next_value * next_non_terminal - values_t[t]
            gae = delta + self.gamma * self.gae_lambda * next_non_terminal * gae
            advantages[t] = gae
        
        returns = advantages + values_t
        advantages = (advantages - advantages.mean()) / (advantages.std() + 1e-8)
        
        return advantages, returns
    
    def get_mini_batches(self, batch_size: int) -> List[np.ndarray]:
        """生成mini-batch索引"""
        indices = np.random.permutation(len(self.transitions))
        batches = []
        for start in range(0, len(self.transitions), batch_size):
            batches.append(indices[start:start + batch_size])
        return batches
    
    @property
    def reward_var(self):
        return getattr(self, '_reward_var', 1.0)
    
    @reward_var.setter
    def reward_var(self, value):
        self._reward_var = value


# ============================================================================
#                              指标计算器 (严格按照定义)
# ============================================================================

class MetricsCalculator:
    """训练指标计算器
    
    严格按照定义计算指标：
    
    农学指标：
    1. 最终产量：决策周期结束时由gym-dssat环境直接输出 (kg/ha)
    2. 灌溉量：决策周期内累计灌溉水量 (mm)
    3. 施肥量：决策周期内累计施氮量 (kg/ha)
    4. WUE：最终产量 / 累计灌溉量 (kg/mm)
    5. NUE：最终产量 / 累计施氮量 (kg/kg)
    
    AI指标：
    6. 平均回报：E[Σ γ^t * r_t] - 折扣累积奖励的期望
    7. 样本效率：达到预设目标性能所需的交互步数
    8. 收敛速度：性能稳定收敛所需的总交互步数
    """
    
    def __init__(self, gamma: float = 0.99, target_performance: float = 800.0,
                 convergence_window: int = 100, convergence_threshold: float = 0.05):
        self.gamma = gamma
        self.target_performance = target_performance
        self.convergence_window = convergence_window
        self.convergence_threshold = convergence_threshold
        self.reset()
        
    def reset(self):
        """重置所有指标"""
        # ========== 农学指标 ==========
        # 1. 最终产量 (kg/ha)
        self.final_yields = []
        # 2. 灌溉量 (mm)
        self.total_irrigations = []
        # 3. 施肥量 (kg N/ha)
        self.total_fertilizers = []
        # 4. WUE (kg/mm) - 水分利用效率
        self.wue_values = []
        # 5. NUE (kg/kg) - 氮肥利用效率
        self.nue_values = []
        
        # ========== AI指标 ==========
        # 6. 平均回报相关（折扣累积奖励）
        self.discounted_returns = []  # 每轮的折扣累积奖励
        self.episode_rewards = []     # 每轮的累积奖励（非折扣）
        self.episode_lengths = []     # 每轮的步数
        
        # 7. 样本效率
        self.total_steps = 0                    # 总交互步数
        self.sample_efficiency_steps = None     # 达到目标性能的步数
        self.reached_target = False             # 是否已达到目标
        
        # 8. 收敛速度
        self.convergence_steps = None           # 收敛时的总交互步数
        self.is_converged = False               # 是否已收敛
        self.return_window = deque(maxlen=self.convergence_window)  # 滑动窗口
        
        # ========== 其他追踪 ==========
        self.policy_losses = []
        self.value_losses = []
        self.entropies = []
        
        self.start_time = time.time()
        self.best_avg_return = float('-inf')
        
    def add_episode(self, rewards: List[float], yield_val: float, 
                    irrigation: float, fertilizer: float,
                    episode_length: int):
        """
        添加一轮训练数据
        
        Args:
            rewards: 该轮所有步骤的即时奖励列表
            yield_val: 最终产量 (kg/ha)
            irrigation: 累计灌溉量 (mm)
            fertilizer: 累计施氮量 (kg N/ha)
            episode_length: 该轮步数
        """
        # ========== 1. 最终产量 ==========
        # 定义：决策周期结束时由gym-dssat环境直接输出
        self.final_yields.append(yield_val)
        
        # ========== 2. 灌溉量 ==========
        # 定义：决策周期内累计灌溉水量
        self.total_irrigations.append(irrigation)
        
        # ========== 3. 施肥量 ==========
        # 定义：决策周期内累计施氮量
        self.total_fertilizers.append(fertilizer)
        
        # ========== 4. WUE (水分利用效率) ==========
        # 定义：最终产量 / 累计灌溉量 (kg/mm)
        # 若灌溉量为0，则无法计算，忽略
        if irrigation > 0:
            wue = yield_val / irrigation
            self.wue_values.append(wue)
        
        # ========== 5. NUE (氮肥利用效率) ==========
        # 定义：最终产量 / 累计施氮量 (kg/kg)
        # 若施肥量为0，则无法计算，忽略
        if fertilizer > 0:
            nue = yield_val / fertilizer
            self.nue_values.append(nue)
        
        # ========== 6. 平均回报 ==========
        # 定义：E[Σ γ^t * r_t] - 折扣累积奖励的期望
        # 计算折扣累积奖励
        discounted_return = 0.0
        for t, r in enumerate(rewards):
            discounted_return += (self.gamma ** t) * r
        self.discounted_returns.append(discounted_return)
        
        # 记录非折扣累积奖励和轮次长度
        self.episode_rewards.append(sum(rewards))
        self.episode_lengths.append(episode_length)
        
        # ========== 7. 样本效率 ==========
        # 定义：达到预设目标性能所需的交互步数
        self.total_steps += episode_length
        
        if not self.reached_target:
            # 使用折扣累积奖励判断是否达到目标性能
            current_avg_return = np.mean(self.discounted_returns[-50:]) if len(self.discounted_returns) >= 50 else np.mean(self.discounted_returns)
            if current_avg_return >= self.target_performance:
                self.sample_efficiency_steps = self.total_steps
                self.reached_target = True
        
        # ========== 8. 收敛速度 ==========
        # 定义：性能稳定收敛所需的训练进程（总交互步数）
        # 判定条件：平均回报稳定在特定阈值范围内不再显著上升
        if not self.is_converged:
            self.return_window.append(discounted_return)
            
            if len(self.return_window) >= self.convergence_window:
                window_arr = np.array(list(self.return_window))
                window_mean = window_arr.mean()
                window_std = window_arr.std()
                
                # 判断收敛：标准差/均值的比例小于阈值，且均值不再显著上升
                if window_mean > 0:
                    cv = window_std / window_mean  # 变异系数
                    if cv < self.convergence_threshold:
                        # 检查是否稳定（最近50步均值与窗口均值接近）
                        recent_mean = np.mean(list(self.return_window)[-50:])
                        if abs(recent_mean - window_mean) / window_mean < self.convergence_threshold:
                            self.convergence_steps = self.total_steps
                            self.is_converged = True
        
        # 更新最佳平均回报
        current_avg = np.mean(self.discounted_returns) if self.discounted_returns else 0
        if current_avg > self.best_avg_return:
            self.best_avg_return = current_avg
    
    def add_update(self, policy_loss: float, value_loss: float, entropy: float):
        """添加更新数据"""
        self.policy_losses.append(policy_loss)
        self.value_losses.append(value_loss)
        self.entropies.append(entropy)
    
    def get_agricultural_metrics(self, last_n: int = None) -> Dict[str, Any]:
        """
        获取农学指标
        
        Returns:
            包含以下指标的字典：
            - yield: 最终产量统计 (kg/ha)
            - irrigation: 灌溉量统计 (mm)
            - fertilizer: 施肥量统计 (kg N/ha)
            - WUE: 水分利用效率统计 (kg/mm)
            - NUE: 氮肥利用效率统计 (kg/kg)
        """
        def get_stats(data, name=""):
            if not data:
                return {'mean': None, 'std': None, 'max': None, 'min': None, 'count': 0}
            arr = np.array(data)
            return {
                'mean': float(arr.mean()),
                'std': float(arr.std()),
                'max': float(arr.max()),
                'min': float(arr.min()),
                'count': len(arr)
            }
        
        if last_n:
            yields = self.final_yields[-last_n:]
            irrigation = self.total_irrigations[-last_n:]
            fertilizer = self.total_fertilizers[-last_n:]
            wue = self.wue_values[-last_n:] if self.wue_values else []
            nue = self.nue_values[-last_n:] if self.nue_values else []
        else:
            yields = self.final_yields
            irrigation = self.total_irrigations
            fertilizer = self.total_fertilizers
            wue = self.wue_values
            nue = self.nue_values
        
        return {
            'yield': get_stats(yields, 'yield'),           # 最终产量 (kg/ha)
            'irrigation': get_stats(irrigation, 'irrigation'),  # 灌溉量 (mm)
            'fertilizer': get_stats(fertilizer, 'fertilizer'),  # 施肥量 (kg N/ha)
            'WUE': get_stats(wue, 'WUE'),                  # 水分利用效率 (kg/mm)
            'NUE': get_stats(nue, 'NUE')                   # 氮肥利用效率 (kg/kg)
        }
    
    def get_ai_metrics(self) -> Dict[str, Any]:
        """
        获取AI指标
        
        Returns:
            包含以下指标的字典：
            - avg_return: 平均回报（折扣累积奖励的期望）
            - sample_efficiency: 样本效率（步）
            - convergence_speed: 收敛速度（步）
            - total_steps: 总交互步数
            - training_time: 训练时间
        """
        # 6. 平均回报：折扣累积奖励的期望
        avg_return = np.mean(self.discounted_returns) if self.discounted_returns else None
        
        # 7. 样本效率：达到预设目标性能所需的交互步数
        sample_efficiency = self.sample_efficiency_steps  # 可能为None（未达到目标）
        
        # 8. 收敛速度：性能稳定收敛所需的总交互步数
        convergence_speed = self.convergence_steps  # 可能为None（未收敛）
        
        training_time = time.time() - self.start_time
        
        # 计算最近N轮的平均回报
        recent_avg_return = np.mean(self.discounted_returns[-100:]) if len(self.discounted_returns) >= 100 else avg_return
        
        return {
            'avg_return': avg_return,                              # 平均回报（无量纲）
            'sample_efficiency': sample_efficiency,                 # 样本效率（步）
            'convergence_speed': convergence_speed,                 # 收敛速度（步）
            'total_steps': self.total_steps,                        # 总交互步数
            'training_time': training_time,                          # 训练时间（秒）
            'best_avg_return': self.best_avg_return,                 # 最佳平均回报
            'recent_avg_return': recent_avg_return,                  # 最近100轮平均回报
            'total_episodes': len(self.discounted_returns),          # 总轮数
            'avg_episode_length': np.mean(self.episode_lengths) if self.episode_lengths else 0,
            'reached_target': self.reached_target,                   # 是否达到目标
            'is_converged': self.is_converged                        # 是否收敛
        }
    
    def get_summary(self, last_n: int = 10) -> str:
        """获取摘要字符串"""
        agri = self.get_agricultural_metrics(last_n)
        ai = self.get_ai_metrics()
        
        # 格式化函数
        def fmt(val, fmt_str="{:.2f}"):
            if val is None:
                return "N/A"
            return fmt_str.format(val)
        
        summary = f"""
┌─────────────────────────────────────────────────────────────────┐
│                    训练指标汇总 (最近{last_n}轮)                    │
├─────────────────────────────────────────────────────────────────┤
│ 农学指标:                                                         │
│   最终产量 (kg/ha):     {fmt(agri['yield']['mean'], '{:.1f}')} ± {fmt(agri['yield']['std'], '{:.1f}')}          │
│   灌溉量 (mm):          {fmt(agri['irrigation']['mean'], '{:.1f}')} ± {fmt(agri['irrigation']['std'], '{:.1f}')}          │
│   施肥量 (kg N/ha):     {fmt(agri['fertilizer']['mean'], '{:.1f}')} ± {fmt(agri['fertilizer']['std'], '{:.1f}')}          │
│   WUE (kg/mm):          {fmt(agri['WUE']['mean'], '{:.2f}')} ± {fmt(agri['WUE']['std'], '{:.2f}')}          │
│   NUE (kg/kg):          {fmt(agri['NUE']['mean'], '{:.2f}')} ± {fmt(agri['NUE']['std'], '{:.2f}')}          │
├─────────────────────────────────────────────────────────────────┤
│ AI指标:                                                           │
│   平均回报:             {fmt(ai['avg_return'], '{:.2f}')}                              │
│   样本效率 (步):        {ai['sample_efficiency'] if ai['sample_efficiency'] else '未达到目标'}                          │
│   收敛速度 (步):        {ai['convergence_speed'] if ai['convergence_speed'] else '未收敛'}                          │
│   总交互步数:           {ai['total_steps']}                                │
│   训练时间:             {ai['training_time']:.1f}s                           │
└─────────────────────────────────────────────────────────────────┘
"""
        return summary
    
    def get_detailed_metrics(self, last_n: int = 10) -> Dict[str, Any]:
        """获取详细指标（用于保存）"""
        return {
            'agricultural': self.get_agricultural_metrics(last_n),
            'ai': self.get_ai_metrics(),
            'timestamp': time.strftime('%Y-%m-%d %H:%M:%S')
        }


# ============================================================================
#                              PPO Agent
# ============================================================================

class PPOAgent:
    """PPO智能体
    
    核心特点：
    1. 作为A2C家族的改进版，使用裁剪目标函数提高稳定性
    2. 使用GAE计算优势估计，平衡偏差和方差
    3. 支持连续学习率调度和早停
    4. 优化的网络初始化和梯度处理
    """
    
    def __init__(self, state_size: int, action_size: int, config: PPOConfig):
        self.config = config
        self.device = device
        
        # 网络
        self.network = ActorCriticNetwork(
            state_size, action_size, config.hidden_sizes
        ).to(self.device)
        
        # 优化器
        self.optimizer = optim.AdamW(
            self.network.parameters(),
            lr=config.learning_rate,
            betas=config.betas,
            eps=config.eps,
            weight_decay=config.weight_decay
        )
        
        # 学习率调度器
        if config.use_lr_schedule:
            total_steps = config.n_episodes * config.max_steps_per_episode // config.update_frequency
            self.lr_scheduler = CosineAnnealingWarmup(
                self.optimizer, config.warmup_steps, total_steps, config.min_lr_ratio
            )
        else:
            self.lr_scheduler = None
        
        # Buffer
        self.buffer = RolloutBuffer(config.gamma, config.gae_lambda, self.device)
        
        # 指标计算器
        self.metrics = MetricsCalculator(
            gamma=config.gamma,
            target_performance=config.target_performance,
            convergence_window=config.convergence_window,
            convergence_threshold=config.convergence_threshold
        )
        
        # 熵系数衰减
        self.current_entropy_coef = config.entropy_coef
        
        # 状态归一化统计量
        self.state_mean = np.zeros(state_size)
        self.state_std = np.ones(state_size)
        self.state_count = 0
        
    def normalize_state(self, state: np.ndarray, update_stats: bool = True) -> np.ndarray:
        """状态归一化"""
        if update_stats:
            self.state_count += 1
            delta = state - self.state_mean
            self.state_mean += delta / self.state_count
            delta2 = state - self.state_mean
            self.state_std = np.sqrt(
                (self.state_std**2 * (self.state_count - 1) + delta * delta2) / self.state_count + 1e-8
            )
        
        return (state - self.state_mean) / (self.state_std + 1e-8)
    
    def act(self, state: np.ndarray, deterministic: bool = False) -> Tuple[int, float, float]:
        """选择动作"""
        self.network.eval()
        
        with torch.no_grad():
            state_t = torch.FloatTensor(state).unsqueeze(0).to(self.device)
            
            with torch.cuda.amp.autocast(enabled=self.config.use_bf16):
                dist, value = self.network.get_action_value(state_t)
            
            if deterministic:
                action = torch.argmax(dist.probs, dim=-1)
            else:
                action = dist.sample()
            
            log_prob = dist.log_prob(action)
        
        return action.item(), log_prob.item(), value.item()
    
    def update(self, last_value: float) -> Dict[str, float]:
        """执行PPO更新"""
        self.network.train()
        
        advantages, returns = self.buffer.compute_gae(last_value, use_reward_norm=True)
        
        all_states = torch.FloatTensor(
            np.array([t.state for t in self.buffer.transitions])
        ).to(self.device)
        all_actions = torch.LongTensor(
            [t.action for t in self.buffer.transitions]
        ).to(self.device)
        all_old_log_probs = torch.FloatTensor(
            [t.log_prob for t in self.buffer.transitions]
        ).to(self.device)
        
        total_policy_loss = 0
        total_value_loss = 0
        total_entropy = 0
        total_kl = 0
        n_updates = 0
        
        for epoch in range(self.config.ppo_epochs):
            mini_batches = self.buffer.get_mini_batches(self.config.mini_batch_size)
            
            for indices in mini_batches:
                self.optimizer.zero_grad()
                
                batch_states = all_states[indices]
                batch_actions = all_actions[indices]
                batch_old_log_probs = all_old_log_probs[indices]
                batch_advantages = advantages[indices]
                batch_returns = returns[indices]
                
                with torch.cuda.amp.autocast(enabled=self.config.use_bf16):
                    dist, new_values = self.network.get_action_value(batch_states)
                    
                    new_log_probs = dist.log_prob(batch_actions)
                    entropy = dist.entropy().mean()
                    
                    ratio = torch.exp(new_log_probs - batch_old_log_probs)
                    
                    surr1 = ratio * batch_advantages
                    surr2 = torch.clamp(ratio, 1 - self.config.clip_ratio, 
                                       1 + self.config.clip_ratio) * batch_advantages
                    policy_loss = -torch.min(surr1, surr2).mean()
                    
                    value_loss = F.mse_loss(new_values.squeeze(), batch_returns)
                    
                    loss = (policy_loss + 
                           self.config.value_coef * value_loss - 
                           self.current_entropy_coef * entropy)
                
                loss.backward()
                
                torch.nn.utils.clip_grad_norm_(
                    self.network.parameters(), 
                    self.config.max_grad_norm
                )
                
                self.optimizer.step()
                
                if self.lr_scheduler:
                    self.lr_scheduler.step()
                
                total_policy_loss += policy_loss.item()
                total_value_loss += value_loss.item()
                total_entropy += entropy.item()
                
                with torch.no_grad():
                    kl = (batch_old_log_probs - new_log_probs).mean().item()
                total_kl += kl
                n_updates += 1
                
                if kl > self.config.target_kl * 1.5:
                    break
            
            if kl > self.config.target_kl * 1.5:
                break
        
        self.buffer.clear()
        
        self.current_entropy_coef = max(
            self.config.entropy_min,
            self.current_entropy_coef * self.config.entropy_decay
        )
        
        return {
            'policy_loss': total_policy_loss / n_updates,
            'value_loss': total_value_loss / n_updates,
            'entropy': total_entropy / n_updates,
            'kl_divergence': total_kl / n_updates,
            'entropy_coef': self.current_entropy_coef
        }
    
    def save(self, path: str, episode: int, extra_info: Dict = None):
        """保存模型"""
        os.makedirs(path, exist_ok=True)
        save_dict = {
            'episode': episode,
            'network': self.network.state_dict(),
            'optimizer': self.optimizer.state_dict(),
            'state_mean': self.state_mean,
            'state_std': self.state_std,
            'state_count': self.state_count,
            'current_entropy_coef': self.current_entropy_coef,
            'config': self.config.__dict__,
        }
        if extra_info:
            save_dict.update(extra_info)
        
        torch.save(save_dict, os.path.join(path, f'model_ep{episode}.pth'))
        
    def load(self, path: str):
        """加载模型"""
        checkpoint = torch.load(path, map_location=self.device)
        self.network.load_state_dict(checkpoint['network'])
        self.optimizer.load_state_dict(checkpoint['optimizer'])
        self.state_mean = checkpoint.get('state_mean', np.zeros(self.config.state_size))
        self.state_std = checkpoint.get('state_std', np.ones(self.config.state_size))
        self.state_count = checkpoint.get('state_count', 0)
        self.current_entropy_coef = checkpoint.get('current_entropy_coef', self.config.entropy_coef)
        return checkpoint


# ============================================================================
#                              动作映射
# ============================================================================

def action_to_dict(action: int, state: np.ndarray) -> Dict[str, float]:
    """将离散动作转换为gym-dssat动作字典
    
    动作空间：25个离散动作 (5氮肥等级 × 5灌溉等级)
    - 氮肥：0, 40, 80, 120, 160 kg/ha
    - 灌溉：0, 6, 12, 18, 24 mm
    """
    n_level = action % 5  # 0-4
    w_level = action // 5  # 0-4
    
    anfer = n_level * 40  # 氮肥量 kg/ha
    amir = w_level * 6    # 灌溉量 mm
    
    # 约束条件
    if state[0] >= 10000:
        anfer = 0
    if state[21] >= 1600:
        amir = 0
    
    return {'anfer': anfer, 'amir': amir}


# ============================================================================
#                              训练函数
# ============================================================================

def save_metrics_to_excel(metrics: Dict[str, Any], filepath: str):
    """将训练指标保存为Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Metrics"
    
    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入农学指标
    ws['A1'] = "Agronomic Metrics"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:E1')
    
    agri_headers = ['Metric', 'Mean', 'Std', 'Max', 'Min']
    for col, header in enumerate(agri_headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    agri_metrics = metrics['agricultural']
    agri_data = [
        ['Yield (kg/ha)', agri_metrics['yield']['mean'], agri_metrics['yield']['std'], 
         agri_metrics['yield']['max'], agri_metrics['yield']['min']],
        ['Irrigation (mm)', agri_metrics['irrigation']['mean'], agri_metrics['irrigation']['std'],
         agri_metrics['irrigation']['max'], agri_metrics['irrigation']['min']],
        ['Fertilizer (kg/ha)', agri_metrics['fertilizer']['mean'], agri_metrics['fertilizer']['std'],
         agri_metrics['fertilizer']['max'], agri_metrics['fertilizer']['min']],
        ['WUE (kg/mm)', agri_metrics['WUE']['mean'], agri_metrics['WUE']['std'],
         agri_metrics['WUE']['max'], agri_metrics['WUE']['min']],
        ['NUE (kg/kg)', agri_metrics['NUE']['mean'], agri_metrics['NUE']['std'],
         agri_metrics['NUE']['max'], agri_metrics['NUE']['min']]
    ]
    
    for row_idx, row_data in enumerate(agri_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value if value is not None else "N/A"
            cell.alignment = center_alignment
    
    # 写入AI指标
    ws['A9'] = "AI Metrics"
    ws['A9'].font = header_font
    ws['A9'].fill = header_fill
    ws.merge_cells('A9:E9')
    
    ai_headers = ['Metric', 'Value']
    for col, header in enumerate(ai_headers, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    ai_metrics = metrics['ai']
    ai_data = [
        ['Average Return', ai_metrics['avg_return']],
        ['Sample Efficiency (steps)', ai_metrics['sample_efficiency'] if ai_metrics['sample_efficiency'] else 'Not reached'],
        ['Convergence Speed (steps)', ai_metrics['convergence_speed'] if ai_metrics['convergence_speed'] else 'Not converged'],
        ['Total Steps', ai_metrics['total_steps']],
        ['Training Time (s)', f"{ai_metrics['training_time']:.1f}"],
        ['Best Average Return', ai_metrics['best_avg_return']],
        ['Recent Average Return', ai_metrics['recent_avg_return']],
        ['Total Episodes', ai_metrics['total_episodes']],
        ['Average Episode Length', f"{ai_metrics['avg_episode_length']:.1f}"],
        ['Reached Target', 'Yes' if ai_metrics['reached_target'] else 'No'],
        ['Is Converged', 'Yes' if ai_metrics['is_converged'] else 'No']
    ]
    
    for row_idx, row_data in enumerate(ai_data, 11):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = center_alignment
    
    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    wb.save(filepath)
    print(f"✓ Metrics saved to Excel: {filepath}")

def plot_training_curves(metrics_calculator: MetricsCalculator):
    """绘制训练曲线图"""
    print("\n[5] Plotting training curves...")
    
    fig, axes = plt.subplots(3, 3, figsize=(15, 12))
    
    window = min(50, len(metrics_calculator.discounted_returns) // 5) if len(metrics_calculator.discounted_returns) > 5 else 1
    
    # 1. Discounted return curve
    ax = axes[0, 0]
    ax.plot(metrics_calculator.discounted_returns, alpha=0.6, label='Discounted Return')
    if window > 1:
        ma = np.convolve(metrics_calculator.discounted_returns, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.discounted_returns)), ma, 'r-', linewidth=2, label=f'Moving Avg({window})')
    ax.set_xlabel('Episode')
    ax.set_ylabel('Return (Discounted)')
    ax.set_title('Average Return Curve')
    ax.legend()
    ax.grid(True, alpha=0.3)
    
    # 2. Yield curve
    ax = axes[0, 1]
    ax.plot(metrics_calculator.final_yields, alpha=0.6)
    if window > 1:
        ma = np.convolve(metrics_calculator.final_yields, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.final_yields)), ma, 'r-', linewidth=2)
    ax.set_xlabel('Episode')
    ax.set_ylabel('Yield (kg/ha)')
    ax.set_title('Final Yield Curve')
    ax.grid(True, alpha=0.3)
    
    # 3. Entropy curve
    ax = axes[0, 2]
    ax.plot(metrics_calculator.entropies, alpha=0.6)
    if window > 1:
        ma = np.convolve(metrics_calculator.entropies, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.entropies)), ma, 'r-', linewidth=2)
    ax.set_xlabel('Episode')
    ax.set_ylabel('Entropy')
    ax.set_title('Policy Entropy')
    ax.grid(True, alpha=0.3)
    
    # 4. Fertilizer amount
    ax = axes[1, 0]
    ax.plot(metrics_calculator.total_fertilizers, alpha=0.6)
    if window > 1:
        ma = np.convolve(metrics_calculator.total_fertilizers, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.total_fertilizers)), ma, 'r-', linewidth=2)
    ax.set_xlabel('Episode')
    ax.set_ylabel('Fertilizer (kg/ha)')
    ax.set_title('Fertilizer Application')
    ax.grid(True, alpha=0.3)
    
    # 5. Irrigation amount
    ax = axes[1, 1]
    ax.plot(metrics_calculator.total_irrigations, alpha=0.6)
    if window > 1:
        ma = np.convolve(metrics_calculator.total_irrigations, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.total_irrigations)), ma, 'r-', linewidth=2)
    ax.set_xlabel('Episode')
    ax.set_ylabel('Irrigation (mm)')
    ax.set_title('Irrigation Application')
    ax.grid(True, alpha=0.3)
    
    # 6. WUE
    ax = axes[1, 2]
    ax.plot(metrics_calculator.wue_values, alpha=0.6)
    if window > 1:
        ma = np.convolve(metrics_calculator.wue_values, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.wue_values)), ma, 'r-', linewidth=2)
    ax.set_xlabel('Episode')
    ax.set_ylabel('WUE (kg/mm)')
    ax.set_title('Water Use Efficiency (WUE)')
    ax.grid(True, alpha=0.3)
    
    # 7. NUE
    ax = axes[2, 0]
    ax.plot(metrics_calculator.nue_values, alpha=0.6)
    if window > 1:
        ma = np.convolve(metrics_calculator.nue_values, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.nue_values)), ma, 'r-', linewidth=2)
    ax.set_xlabel('Episode')
    ax.set_ylabel('NUE (kg/kg)')
    ax.set_title('Nitrogen Use Efficiency (NUE)')
    ax.grid(True, alpha=0.3)
    
    # 8. Policy loss and value loss
    ax = axes[2, 1]
    ax.plot(metrics_calculator.policy_losses, alpha=0.6, label='Policy Loss')
    ax.plot(metrics_calculator.value_losses, alpha=0.6, label='Value Loss')
    if window > 1:
        ma_policy = np.convolve(metrics_calculator.policy_losses, np.ones(window)/window, mode='valid')
        ma_value = np.convolve(metrics_calculator.value_losses, np.ones(window)/window, mode='valid')
        ax.plot(range(window-1, len(metrics_calculator.policy_losses)), ma_policy, 'r-', linewidth=2, label='Policy MA')
        ax.plot(range(window-1, len(metrics_calculator.value_losses)), ma_value, 'b-', linewidth=2, label='Value MA')
    ax.set_xlabel('Update')
    ax.set_ylabel('Loss')
    ax.set_title('Training Losses')
    ax.legend()
    ax.grid(True, alpha=0.3)
    
    # 9. Metrics summary table
    ax = axes[2, 2]
    ax.axis('off')
    
    agri_metrics = metrics_calculator.get_agricultural_metrics()
    ai_metrics = metrics_calculator.get_ai_metrics()
    
    summary_text = """
    [Final Metrics Summary]
    
    Agronomic Metrics:
    ─────────────────────
    Avg Yield: {:.2f} kg/ha
    Max Yield: {:.2f} kg/ha
    Avg WUE: {:.4f} kg/mm
    Avg NUE: {:.2f} kg/kg
    
    AI Metrics:
    ─────────────────────
    Avg Return: {:.2f}
    Best Return: {:.2f}
    Sample Efficiency: {} steps
    Convergence Speed: {} steps
    Total Episodes: {}
    """.format(
        agri_metrics['yield']['mean'],
        agri_metrics['yield']['max'],
        agri_metrics['WUE']['mean'],
        agri_metrics['NUE']['mean'],
        ai_metrics['avg_return'],
        ai_metrics['best_avg_return'],
        ai_metrics['sample_efficiency'] if ai_metrics['sample_efficiency'] else 'Not reached',
        ai_metrics['convergence_speed'] if ai_metrics['convergence_speed'] else 'Not converged',
        ai_metrics['total_episodes']
    )
    
    ax.text(0.1, 0.5, summary_text, transform=ax.transAxes, fontsize=10,
            verticalalignment='center', fontfamily='monospace',
            bbox=dict(boxstyle='round', facecolor='lightgray', alpha=0.5))
    
    plt.tight_layout()
    
    # Save figures
    plot_path = os.path.join(RESULTS_DIR, 'ppo_baseline_training_curves.pdf')
    plt.savefig(plot_path, dpi=150, bbox_inches='tight')
    print(f"✓ Training curves saved: {plot_path}")
    
    png_path = os.path.join(RESULTS_DIR, 'ppo_baseline_training_curves.png')
    plt.savefig(png_path, dpi=150, bbox_inches='tight')
    print(f"✓ Training curves saved: {png_path}")
    
    plt.close()

def train_ppo():
    """PPO训练主函数"""
    print("\n" + "=" * 70)
    print("🚀 开始PPO训练")
    print("=" * 70)
    
    # 设置随机种子
    set_seed(42)
    
    # 创建环境
    print("\n📦 初始化环境...")
    env_args = {
        'run_dssat_location': '/opt/dssat_pdi/run_dssat',
        'log_saving_path': '/home/wuyang/data/logs/dssat-pdi.log',
        'mode': 'all',
        'seed': 123456,
        'random_weather': True
    }
    
    try:
        env = gym.make('gym_dssat_pdi:GymDssatPdi-v0', **env_args)
        print("   ✓ 环境创建成功")
    except Exception as e:
        print(f"   ✗ 环境创建失败: {e}")
        print("   请确保gym-dssat已正确安装")
        return
    
    # 获取状态维度
    sample_state = dict2array(env.reset())
    state_size = len(sample_state)
    print(f"   状态维度: {state_size}")
    
    # 更新配置
    config.state_size = state_size
    
    # 创建智能体
    print("\n🤖 创建PPO智能体...")
    agent = PPOAgent(state_size, config.action_size, config)
    print(f"   网络参数量: {sum(p.numel() for p in agent.network.parameters()):,}")
    
    # 创建保存目录
    os.makedirs('/home/wuyang/checkpoints/only_ppo_checkpoints', exist_ok=True)
    os.makedirs('/home/wuyang/data/logs', exist_ok=True)
    
    # 训练记录
    all_metrics = {
        'episodes': [],
        'agricultural': [],
        'ai': []
    }
    
    print(f"\n🎯 开始训练 ({config.n_episodes} 轮)")
    print(f"   目标性能阈值: {config.target_performance}")
    print("-" * 70)
    
    best_avg_return = float('-inf')
    
    # 创建进度条
    pbar = tqdm(total=config.n_episodes, desc="Training Progress", 
                unit="ep", ncols=100, position=0, leave=True)
    
    # 记录训练开始时间
    training_start_time = time.time()
    
    for episode in range(1, config.n_episodes + 1):
        # 重置环境
        state_raw = dict2array(env.reset())
        state = agent.normalize_state(state_raw, update_stats=True)
        
        # 轮次追踪
        episode_rewards = []  # 存储每步的即时奖励
        episode_n_amount = 0  # 累计氮肥量
        episode_w_amount = 0  # 累计灌溉量
        episode_yield = 0
        episode_length = 0
        
        for step in range(config.max_steps_per_episode):
            # 选择动作
            action, log_prob, value = agent.act(state)
            
            # 转换为环境动作
            action_dict = action_to_dict(action, state_raw)
            
            # 执行动作
            next_state_raw, _, done, _ = env.step(action_dict)
            next_state_raw = dict2array(next_state_raw) if not done else state_raw
            
            # 计算奖励
            reward = get_reward(
                state_raw, action_dict['anfer'], action_dict['amir'],
                next_state_raw, done, config.k1, config.k2, config.k3
            )
            
            # 存入buffer
            agent.buffer.add(
                state, action, reward, 
                agent.normalize_state(next_state_raw, update_stats=False),
                done, log_prob, value,
                action_dict['anfer'], action_dict['amir']
            )
            
            # 记录即时奖励
            episode_rewards.append(reward)
            
            # 更新状态
            state_raw = next_state_raw
            state = agent.normalize_state(state_raw, update_stats=True)
            
            # 累计资源使用
            episode_n_amount += action_dict['anfer']
            episode_w_amount += action_dict['amir']
            episode_length += 1
            
            if done:
                # 最终产量：环境直接输出
                episode_yield = state_raw[4]
                break
        
        # 记录轮次数据到指标计算器
        agent.metrics.add_episode(
            rewards=episode_rewards,      # 所有即时奖励
            yield_val=episode_yield,       # 最终产量
            irrigation=episode_w_amount,   # 累计灌溉量
            fertilizer=episode_n_amount,   # 累计施肥量
            episode_length=episode_length  # 轮次长度
        )
        
        # 更新策略
        if episode % config.update_frequency == 0 and len(agent.buffer) >= config.mini_batch_size:
            with torch.no_grad():
                last_value = agent.act(state, deterministic=True)[2]
                if done:
                    last_value = 0.0
            
            update_info = agent.update(last_value)
            agent.metrics.add_update(
                update_info['policy_loss'],
                update_info['value_loss'],
                update_info['entropy']
            )
        
        # 更新进度条
        if episode % config.log_frequency == 0:
            agri_metrics = agent.metrics.get_agricultural_metrics(last_n=10)
            ai_metrics = agent.metrics.get_ai_metrics()
            lr = agent.lr_scheduler.get_lr()[0] if agent.lr_scheduler else config.learning_rate
            
            # 格式化输出
            wue_str = f"{agri_metrics['WUE']['mean']:.1f}" if agri_metrics['WUE']['mean'] is not None else "N/A"
            nue_str = f"{agri_metrics['NUE']['mean']:.1f}" if agri_metrics['NUE']['mean'] is not None else "N/A"
            se_str = f"{ai_metrics['sample_efficiency']}" if ai_metrics['sample_efficiency'] else "---"
            
            # 更新进度条描述
            pbar.set_postfix_str(
                f"Return: {ai_metrics['avg_return']:.1f} | "
                f"Yield: {episode_yield:.0f} | "
                f"N: {episode_n_amount:.0f} | W: {episode_w_amount:.0f} | "
                f"WUE: {wue_str} | NUE: {nue_str} | "
                f"Steps: {ai_metrics['total_steps']} | SE: {se_str}"
            )
        
        # 保存最佳模型
        current_avg = np.mean(agent.metrics.discounted_returns[-100:]) if len(agent.metrics.discounted_returns) >= 100 else np.mean(agent.metrics.discounted_returns)
        if current_avg > best_avg_return and current_avg>1400:
            best_avg_return = current_avg
            agent.save('/home/wuyang/checkpoints/only_ppo_checkpoints/best', episode, {
                'avg_return': current_avg,
                'best_yield': episode_yield
            })
        
        # 定期保存
        if episode % config.save_frequency == 0:
            agent.save('/home/wuyang/checkpoints/only_ppo_checkpoints', episode)
            
            # 保存指标
            all_metrics['episodes'].append(episode)
            all_metrics['agricultural'].append(agent.metrics.get_agricultural_metrics())
            all_metrics['ai'].append(agent.metrics.get_ai_metrics())
        
        # 更新进度条
        pbar.update(1)
    
    # 关闭进度条
    pbar.close()
    
    # 计算总训练时间
    total_training_time = time.time() - training_start_time
    print(f"\n⏱️  总训练时间: {total_training_time/3600:.2f} 小时 ({total_training_time:.1f} 秒)")
    print(f"⏱️  平均每轮耗时: {total_training_time/config.n_episodes:.2f} 秒")
    
    # 训练完成
    env.close()
    
    print("\n" + "=" * 70)
    print("🎉 训练完成！")
    print("=" * 70)
    
    # 输出最终指标汇总
    print(agent.metrics.get_summary(last_n=10))
    
    # 输出整体训练指标
    print("\n" + "=" * 70)
    print("📊 整体训练指标汇总")
    print("=" * 70)
    print(agent.metrics.get_summary(last_n=len(agent.metrics.final_yields)))
    
    # 保存最终模型和指标
    agent.save('/home/wuyang/checkpoints/only_ppo_checkpoints/final', config.n_episodes)
    
    # 绘制训练曲线
    plot_training_curves(agent.metrics)
    
    # 保存指标到Excel
    detailed_metrics = agent.metrics.get_detailed_metrics()
    excel_path = os.path.join(RESULTS_DIR, 'ppo_baseline_metrics.xlsx')
    save_metrics_to_excel(detailed_metrics, excel_path)

    
    with open('/home/wuyang/results/only_ppo_results/training_metrics.json', 'w') as f:
        def convert(obj):
            if isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, (np.int32, np.int64)):
                return int(obj)
            elif isinstance(obj, (np.float32, np.float64)):
                return float(obj)
            elif obj is None:
                return None
            return obj
        
        json.dump(all_metrics, f, default=convert, indent=2)
    
    return agent


def evaluate_agent(agent: PPOAgent, n_episodes: int = 10):
    """评估智能体"""
    print("\n" + "=" * 70)
    print(f"📊 模型评估 ({n_episodes} 轮)")
    print("=" * 70)
    
    # 创建环境
    env_args = {
        'run_dssat_location': '/opt/dssat_pdi/run_dssat',
        'log_saving_path': '/home/wuyang/data/logs/dssat-pdi-validation.log',
        'mode': 'all',
        'seed': 999,
        'random_weather': True
    }
    
    env = gym.make('gym_dssat_pdi:GymDssatPdi-v0', **env_args)
    
    # 创建评估指标计算器
    eval_metrics = MetricsCalculator(
        gamma=config.gamma,
        target_performance=config.target_performance
    )
    
    for ep in range(n_episodes):
        state_raw = dict2array(env.reset())
        state = agent.normalize_state(state_raw, update_stats=False)
        
        episode_rewards = []
        episode_n_amount = 0
        episode_w_amount = 0
        episode_yield = 0
        episode_length = 0
        
        done = False
        while not done and episode_length < config.max_steps_per_episode:
            action, _, _ = agent.act(state, deterministic=True)
            
            action_dict = action_to_dict(action, state_raw)
            next_state_raw, _, done, _ = env.step(action_dict)
            next_state_raw = dict2array(next_state_raw) if not done else state_raw
            
            reward = get_reward(
                state_raw, action_dict['anfer'], action_dict['amir'],
                next_state_raw, done, config.k1, config.k2, config.k3
            )
            
            episode_rewards.append(reward)
            
            state_raw = next_state_raw if not done else state_raw
            state = agent.normalize_state(state_raw, update_stats=False)
            
            episode_n_amount += action_dict['anfer']
            episode_w_amount += action_dict['amir']
            episode_length += 1
            
            if done:
                episode_yield = state_raw[4]
        
        # 记录评估数据
        eval_metrics.add_episode(
            rewards=episode_rewards,
            yield_val=episode_yield,
            irrigation=episode_w_amount,
            fertilizer=episode_n_amount,
            episode_length=episode_length
        )
        
        print(f"Eval {ep+1:2d} | "
              f"Return: {eval_metrics.discounted_returns[-1]:.1f} | "
              f"Yield: {episode_yield:.0f} kg/ha | "
              f"N: {episode_n_amount:.0f} | W: {episode_w_amount:.0f}")
    
    env.close()
    
    # 输出评估结果
    print("\n" + "=" * 70)
    print("📊 评估结果汇总")
    print("=" * 70)
    print(eval_metrics.get_summary(last_n=n_episodes))
    
    return eval_metrics


# ============================================================================
#                              主程序
# ============================================================================

if __name__ == "__main__":
    # 训练
    agent = train_ppo()
    
    # 评估
    if agent:
        evaluate_agent(agent, n_episodes=config.n_eval_episodes)
    
    print("\n✅ 所有任务完成！")